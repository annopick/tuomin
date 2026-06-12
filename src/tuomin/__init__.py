#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 脱敏工具 - 对 xls/xlsx 文件中的个人敏感信息进行替换或重构脱敏"""

import argparse
import os
import random
import re
import string
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

FIELD_TYPES = {"name", "phone", "idcard", "email", "bankcard", "address"}
METHODS = {"replace", "reconstruct"}

SURNAMES = [
    "赵", "钱", "孙", "李", "周", "吴", "郑", "王", "冯", "陈",
    "褚", "卫", "蒋", "沈", "韩", "杨", "朱", "秦", "尤", "许",
    "何", "吕", "施", "张", "孔", "曹", "严", "华", "金", "魏",
    "陶", "姜", "戚", "谢", "邹", "喻", "柏", "水", "窦", "章",
    "云", "苏", "潘", "葛", "奚", "范", "彭", "郎", "鲁", "韦",
    "昌", "马", "苗", "凤", "花", "方", "俞", "任", "袁", "柳",
    "酆", "鲍", "史", "唐", "费", "廉", "岑", "薛", "雷", "贺",
    "倪", "汤", "滕", "殷", "罗", "毕", "郝", "邬", "安", "常",
    "乐", "于", "时", "傅", "皮", "卞", "齐", "康", "伍", "余",
    "元", "卜", "顾", "孟", "平", "黄", "和", "穆", "萧", "尹",
]

CN_DIGITS = "一二三四五六七八九"

PROVINCES_CITIES = [
    ("北京市", "北京市"),
    ("上海市", "上海市"),
    ("天津市", "天津市"),
    ("重庆市", "重庆市"),
    ("广东省", "广州市"),
    ("广东省", "深圳市"),
    ("广东省", "珠海市"),
    ("广东省", "东莞市"),
    ("浙江省", "杭州市"),
    ("浙江省", "宁波市"),
    ("浙江省", "温州市"),
    ("江苏省", "南京市"),
    ("江苏省", "苏州市"),
    ("江苏省", "无锡市"),
    ("四川省", "成都市"),
    ("四川省", "绵阳市"),
    ("湖北省", "武汉市"),
    ("湖北省", "宜昌市"),
    ("湖南省", "长沙市"),
    ("湖南省", "株洲市"),
    ("福建省", "福州市"),
    ("福建省", "厦门市"),
    ("山东省", "济南市"),
    ("山东省", "青岛市"),
    ("河南省", "郑州市"),
    ("河南省", "洛阳市"),
    ("河北省", "石家庄市"),
    ("河北省", "唐山市"),
    ("辽宁省", "沈阳市"),
    ("辽宁省", "大连市"),
    ("吉林省", "长春市"),
    ("黑龙江省", "哈尔滨市"),
    ("陕西省", "西安市"),
    ("山西省", "太原市"),
    ("安徽省", "合肥市"),
    ("江西省", "南昌市"),
    ("云南省", "昆明市"),
    ("贵州省", "贵阳市"),
    ("甘肃省", "兰州市"),
    ("广西壮族自治区", "南宁市"),
    ("海南省", "海口市"),
    ("内蒙古自治区", "呼和浩特市"),
    ("西藏自治区", "拉萨市"),
    ("宁夏回族自治区", "银川市"),
    ("新疆维吾尔自治区", "乌鲁木齐市"),
    ("青海省", "西宁市"),
]

# 省级关键字（用于地址替换时识别省市边界）
PROVINCE_KEYWORDS = [
    "北京市", "天津市", "上海市", "重庆市",
    "省", "自治区",
]

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------


def col_letter_to_index(letter: str) -> int:
    """Excel 列字母转 0-based 索引。A→0, B→1, Z→25, AA→26"""
    result = 0
    for ch in letter.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"无效的列字母: {letter}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def parse_col_spec(spec: str):
    """解析 --col 参数，格式 'A:phone:replace' → (index, field_type, method)"""
    parts = spec.split(":")
    if len(parts) != 3:
        raise ValueError(f"无效的 --col 格式 '{spec}'，应为 '列字母:字段类型:脱敏方式'")
    col_letter, field_type, method = parts
    col_letter = col_letter.strip().upper()
    field_type = field_type.strip().lower()
    method = method.strip().lower()
    if field_type not in FIELD_TYPES:
        raise ValueError(f"不支持的字段类型 '{field_type}'，可选: {', '.join(sorted(FIELD_TYPES))}")
    if method not in METHODS:
        raise ValueError(f"不支持的脱敏方式 '{method}'，可选: replace, reconstruct")
    idx = col_letter_to_index(col_letter)
    return idx, field_type, method


# ---------------------------------------------------------------------------
# 替换规则
# ---------------------------------------------------------------------------


def replace_name(value: str) -> str:
    if not value:
        return value
    return value[0] + "*" * (len(value) - 1)


def replace_phone(value: str) -> str:
    digits = re.sub(r"\D", "", str(value))
    if len(digits) != 11:
        return str(value)
    return digits[:3] + "****" + digits[7:]


def replace_idcard(value: str) -> str:
    s = str(value).strip().upper()
    if len(s) < 8:
        return str(value)
    return s[:3] + "*" * (len(s) - 7) + s[-4:]


def replace_email(value: str) -> str:
    s = str(value).strip()
    if "@" not in s:
        return s
    local, domain = s.rsplit("@", 1)
    if len(local) <= 1:
        return s
    return local[0] + "*" * (len(local) - 1) + "@" + domain


def replace_bankcard(value: str) -> str:
    digits = re.sub(r"\D", "", str(value))
    if len(digits) < 9:
        return str(value)
    return digits[:4] + "*" * (len(digits) - 8) + digits[-4:]


def replace_address(value: str) -> str:
    s = str(value).strip()
    if not s:
        return s
    # 找到省/市/自治区 结束的位置
    end = 0
    for kw in PROVINCE_KEYWORDS:
        pos = s.find(kw)
        if pos != -1:
            candidate = pos + len(kw)
            if candidate > end:
                end = candidate
    # 再找第一个 "市" 或 "区" 或 "县"
    if end > 0:
        rest = s[end:]
        for target in ["市", "区", "县"]:
            pos = rest.find(target)
            if pos != -1:
                end += pos + len(target)
                break
    if end <= 0:
        return s
    return s[:end] + "*" * (len(s) - end)


REPLACE_FUNCS = {
    "name": replace_name,
    "phone": replace_phone,
    "idcard": replace_idcard,
    "email": replace_email,
    "bankcard": replace_bankcard,
    "address": replace_address,
}


# ---------------------------------------------------------------------------
# 重构规则
# ---------------------------------------------------------------------------


def _rand_cn_digits(length: int) -> str:
    return "".join(random.choice(CN_DIGITS) for _ in range(length))


def reconstruct_name(_value: str) -> str:
    surname = random.choice(SURNAMES)
    name_len = random.randint(2, 3)
    return surname + _rand_cn_digits(name_len)


def reconstruct_phone(_value: str) -> str:
    return "110" + "".join(random.choice(string.digits) for _ in range(8))


def _idcard_check_digit(body: str) -> str:
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_map = "10X98765432"
    total = sum(int(body[i]) * weights[i] for i in range(17))
    return check_map[total % 11]


def reconstruct_idcard(_value: str) -> str:
    region = "999999"
    year = random.randint(1950, 2025)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    seq = random.randint(100, 999)
    body = f"{region}{year:04d}{month:02d}{day:02d}{seq:03d}"
    return body + _idcard_check_digit(body)


def reconstruct_email(_value: str) -> str:
    chars = "".join(random.choice(string.ascii_lowercase + string.digits) for _ in range(5))
    return f"{chars}@notexist.invalid"


def reconstruct_bankcard(_value: str) -> str:
    return "0000" + "".join(random.choice(string.digits) for _ in range(12))


def reconstruct_address(_value: str) -> str:
    prov, city = random.choice(PROVINCES_CITIES)
    road_num = random.randint(100, 999)
    return f"{prov}{city}虚构路{road_num}号"


RECONSTRUCT_FUNCS = {
    "name": reconstruct_name,
    "phone": reconstruct_phone,
    "idcard": reconstruct_idcard,
    "email": reconstruct_email,
    "bankcard": reconstruct_bankcard,
    "address": reconstruct_address,
}


# ---------------------------------------------------------------------------
# 脱敏引擎
# ---------------------------------------------------------------------------


class MaskingEngine:
    def __init__(self):
        # key: (field_type, method) → {原文: 脱敏结果}
        self._mapping: dict[tuple, dict[str, str]] = {}

    def mask(self, value, field_type: str, method: str) -> str:
        """对单个值进行脱敏。返回脱敏后的字符串。"""
        if value is None:
            return value
        s = str(value).strip()
        if not s:
            return value

        key = (field_type, method)
        cache = self._mapping.setdefault(key, {})
        raw_key = s
        if raw_key in cache:
            return cache[raw_key]

        if method == "replace":
            result = REPLACE_FUNCS[field_type](s)
        else:
            result = RECONSTRUCT_FUNCS[field_type](s)

        cache[raw_key] = result
        return result


# ---------------------------------------------------------------------------
# 文件读取
# ---------------------------------------------------------------------------


def read_sheet(filepath: str, sheet=None) -> tuple:
    """读取 Excel 文件的指定 sheet，返回 (rows, col_count)。
    rows 为二维列表，每个元素已转为字符串。
    同时返回用于 openpyxl 复制的原始 workbook 对象（xlsx 时）。
    返回 (rows, col_count, wb_or_none)
    """
    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx(filepath, sheet)
    elif ext == ".xls":
        return _read_xls(filepath, sheet)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xls 和 .xlsx")


def _read_xlsx(filepath: str, sheet=None) -> tuple:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    if sheet is not None:
        if isinstance(sheet, int):
            names = wb.sheetnames
            if sheet < 0 or sheet >= len(names):
                wb.close()
                raise ValueError(f"Sheet 索引 {sheet} 超出范围，共 {len(names)} 个 Sheet")
            ws = wb[names[sheet]]
        else:
            if sheet not in wb.sheetnames:
                wb.close()
                raise ValueError(f"Sheet '{sheet}' 不存在，可选: {', '.join(wb.sheetnames)}")
            ws = wb[sheet]
    else:
        ws = wb.active

    rows = []
    col_count = 0
    for row in ws.iter_rows(values_only=True):
        str_row = []
        for cell in row:
            if cell is None:
                str_row.append("")
            else:
                str_row.append(str(cell))
        rows.append(str_row)
        col_count = max(col_count, len(str_row))

    wb.close()
    return rows, col_count, None


def _read_xls(filepath: str, sheet=None) -> tuple:
    import xlrd

    try:
        wb = xlrd.open_workbook(filepath, encoding_override="gbk")
    except Exception:
        wb = xlrd.open_workbook(filepath)

    if sheet is not None:
        if isinstance(sheet, int):
            if sheet < 0 or sheet >= wb.nsheets:
                raise ValueError(f"Sheet 索引 {sheet} 超出范围，共 {wb.nsheets} 个 Sheet")
            ws = wb.sheet_by_index(sheet)
        else:
            try:
                ws = wb.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                raise ValueError(
                    f"Sheet '{sheet}' 不存在，可选: {', '.join(wb.sheet_names())}"
                )
    else:
        ws = wb.sheet_by_index(0)

    rows = []
    col_count = ws.ncols
    for r in range(ws.nrows):
        str_row = []
        for c in range(col_count):
            cell = ws.cell_value(r, c)
            if cell is None:
                str_row.append("")
            else:
                str_row.append(str(cell))
        rows.append(str_row)

    return rows, col_count, None


# ---------------------------------------------------------------------------
# 文件写入
# ---------------------------------------------------------------------------


def write_xlsx(rows: list, output_path: str):
    """将二维列表写入 xlsx 文件"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# 输出路径
# ---------------------------------------------------------------------------


def resolve_output_path(input_path: str, output_arg=None) -> str:
    if output_arg:
        p = Path(output_arg)
        if p.exists() and p.samefile(Path(input_path)):
            raise ValueError("输出路径不能与输入文件相同")
        return str(p.resolve())

    p = Path(input_path)
    return str(p.with_name(p.stem + "_masked.xlsx").resolve())


# ---------------------------------------------------------------------------
# argparse
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Excel 脱敏工具 - 对 xls/xlsx 文件中的个人敏感信息进行替换或重构脱敏",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
示例:
  python tuomin.py data.xlsx --col C:phone:replace
  python tuomin.py data.xls --sheet Sheet1 --col B:name:reconstruct --col D:idcard:reconstruct --output result.xlsx

支持的字段类型: name, phone, idcard, email, bankcard, address
脱敏方式: replace (用*替换), reconstruct (生成假数据)
""",
    )
    parser.add_argument("input_file", help="输入的 xls/xlsx 文件路径")
    parser.add_argument("--sheet", help="Sheet 名称或索引(从0开始)，默认第一个 Sheet")
    parser.add_argument(
        "--col",
        action="append",
        required=True,
        metavar="COL:TYPE:METHOD",
        help="指定脱敏列，格式如 'A:phone:replace'，可多次使用",
    )
    parser.add_argument("--output", help="输出文件路径，默认在同目录生成 _masked.xlsx")
    return parser


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    parser = build_parser()
    args = parser.parse_args()

    # 验证输入文件
    if not os.path.isfile(args.input_file):
        print(f"错误: 文件不存在: {args.input_file}", file=sys.stderr)
        sys.exit(1)

    # 解析 --col 参数
    col_specs = []
    for spec in args.col:
        try:
            idx, field_type, method = parse_col_spec(spec)
            col_specs.append((idx, field_type, method))
        except ValueError as e:
            print(f"错误: {e}", file=sys.stderr)
            sys.exit(1)

    # 解析 --sheet
    sheet = None
    if args.sheet is not None:
        try:
            sheet = int(args.sheet)
        except ValueError:
            sheet = args.sheet

    # 读取文件
    try:
        rows, col_count, _ = read_sheet(args.input_file, sheet)
    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"错误: 读取文件失败: {e}", file=sys.stderr)
        sys.exit(1)

    # 检查列索引
    for idx, field_type, method in col_specs:
        if idx >= col_count:
            print(
                f"错误: 列索引超出范围，文件共 {col_count} 列，指定列需要至少 {idx + 1} 列",
                file=sys.stderr,
            )
            sys.exit(1)

    # 脱敏
    engine = MaskingEngine()
    for row in rows:
        for idx, field_type, method in col_specs:
            if idx < len(row):
                original = row[idx]
                masked = engine.mask(original, field_type, method)
                row[idx] = masked

    # 输出路径
    try:
        output_path = resolve_output_path(args.input_file, args.output)
    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)

    # 写入
    try:
        write_xlsx(rows, output_path)
    except Exception as e:
        print(f"错误: 写入文件失败: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"脱敏完成，输出文件: {output_path}")


if __name__ == "__main__":
    main()
